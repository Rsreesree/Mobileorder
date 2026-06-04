import webview
import os
import sys
import json
import uuid
import hashlib
import platform
import datetime
import base64
import socket
import struct
import time
import sqlite3
 
from cryptography.hazmat.primitives.asymmetric import padding
from cryptography.hazmat.primitives import hashes, serialization
from cryptography.hazmat.backends import default_backend
 
try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False
 
# ─── paths ────────────────────────────────────────────────────────────────────
def resource_path(filename):
    if hasattr(sys, '_MEIPASS'):
        return os.path.join(sys._MEIPASS, filename)
    return os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
 
# ── Hidden data folder (sits next to the app, invisible in file explorer) ──────
_APP_DIR     = os.path.dirname(os.path.abspath(__file__))
_DATA_DIR    = os.path.join(_APP_DIR, '.hbillsoft')
os.makedirs(_DATA_DIR, exist_ok=True)
 
# On Windows, mark the folder as hidden via attrib
if platform.system() == 'Windows':
    try:
        import subprocess
        subprocess.call(['attrib', '+H', _DATA_DIR], shell=False)
    except Exception:
        pass
 
def data_path(filename):
    """Resolve a data file into the hidden .hbillsoft folder."""
    return os.path.join(_DATA_DIR, filename)
 
CONFIG_FILE    = data_path('config.json')
LICENSE_FILE   = data_path('license.dat')
LASTRUN_FILE   = data_path('lastrun.dat')   # stores last-seen date (hidden from user)
EXCEL_DB_FILE  = data_path('sales_data.xlsx')  # auto-saved sales database
DB_FILE        = data_path('sales_data.db')    # SQLite database
 
# ─── RSA public key (vendor embeds this; private key stays with vendor) ───────
# NOTE: When you generate a new key pair in keygen_gui.py (tkinter),
#       replace the block below with the new public key shown in that dialog.
PUBLIC_KEY_PEM = b"""-----BEGIN PUBLIC KEY-----
MIIBIjANBgkqhkiG9w0BAQEFAAOCAQ8AMIIBCgKCAQEA8N47o0fqzoqxp/bYRokm
sNyT15U7zOc5v/dK7d18ebiblOpTnEpAPEqwg4id/DeAEJyE+oizP8VO2hE21GOo
qTayv6/K9nWOsz17fgzEIpqAAwN8tmeNFrnfOb6UDJ+s4UZPuE3Y8E8hxgTIZWfZ
FA97+tw55p05FwXwxunnQXhGNMPFJppz4iDk9hnZSzIPagj53PMjg/VFcH21I2Yz
8Ey5UWU/P25KnknLWNhwZ3xHrIwJcf29YFxeIE5sEfK8F+4SSx3Te7cI/UwZNWqS
xcMjdssJaEu69EqNbQbYti3SmSAfyNQh1P6VHKVI3jGhLRdvEY7Ml7JWWlfHvD5o
swIDAQAB
-----END PUBLIC KEY-----"""
 
# ─── Clock tamper detection ───────────────────────────────────────────────────
 
# Max allowed drift between system clock and NTP (seconds)
NTP_DRIFT_TOLERANCE = 86400          # 1 day — generous for offline use
# NTP servers to try (in order)
NTP_SERVERS = [
    'time.cloudflare.com',
    'time.google.com',
    'pool.ntp.org',
    'time.windows.com',
]
 
def _get_ntp_time() -> datetime.date | None:
    """
    Query NTP servers for the real current time.
    Returns a date if reachable, None if all servers fail (offline).
    Uses raw UDP — no external libraries needed.
    """
    NTP_DELTA = 2208988800  # seconds between 1900-01-01 and 1970-01-01
    for server in NTP_SERVERS:
        try:
            client = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            client.settimeout(3)
            data = b'\x1b' + 47 * b'\0'
            client.sendto(data, (server, 123))
            response, _ = client.recvfrom(1024)
            client.close()
            # Transmit timestamp is at bytes 40-47
            tx_timestamp = struct.unpack('!I', response[40:44])[0]
            ntp_epoch    = tx_timestamp - NTP_DELTA
            return datetime.date.fromtimestamp(ntp_epoch)
        except Exception:
            continue
    return None  # all servers unreachable (offline)
 
 
def _load_last_run_date() -> datetime.date | None:
    """Load the last-seen date stored on disk (obfuscated, not plain text)."""
    try:
        with open(LASTRUN_FILE, 'r') as f:
            raw = f.read().strip()
        # Decode: base64 → reverse → date string
        decoded = base64.b64decode(raw.encode()).decode()[::-1]
        return datetime.date.fromisoformat(decoded)
    except Exception:
        return None
 
 
def _save_last_run_date(date: datetime.date):
    """Persist today's date to disk (lightly obfuscated)."""
    try:
        # Obfuscate: reverse string → base64 (not encryption, just not plain-text)
        raw     = date.isoformat()[::-1]
        encoded = base64.b64encode(raw.encode()).decode()
        with open(LASTRUN_FILE, 'w') as f:
            f.write(encoded)
    except Exception:
        pass  # read-only filesystem edge case — don't crash
 
 
def detect_clock_tampering(issued_date_str: str) -> dict:
    """
    Run all three clock-tamper checks.
    Returns:
      tampered (bool)  — True if tampering detected
      reason   (str)   — human-readable reason if tampered
      source   (str)   — 'ntp' | 'lastrun' | 'issued' | 'ok'
    """
    today = datetime.date.today()
 
    # ── Layer 1: Issued-date check (always offline) ───────────────────────────
    # If today is BEFORE the issue date, the clock was clearly rolled back.
    try:
        issued = datetime.date.fromisoformat(issued_date_str)
        if today < issued:
            return {
                'tampered': True,
                'reason':   'System date is before the license issue date. Please correct your system clock.',
                'source':   'issued'
            }
    except Exception:
        pass
 
    # ── Layer 2: Last-run date check (always offline) ─────────────────────────
    # If today is BEFORE the last time the app ran, the clock was rolled back.
    last_run = _load_last_run_date()
    if last_run and today < last_run:
        return {
            'tampered': True,
            'reason':   'System date appears to have been set back. Please correct your system clock.',
            'source':   'lastrun'
        }
 
    # ── Layer 3: NTP check (online, best-effort) ──────────────────────────────
    # If we can reach a time server, compare against system clock.
    ntp_date = _get_ntp_time()
    if ntp_date is not None:
        drift_days = abs((today - ntp_date).days)
        if drift_days > 1:   # more than 1 day off
            return {
                'tampered': True,
                'reason':   f'System clock is {drift_days} day(s) off from internet time. Please correct your system clock.',
                'source':   'ntp'
            }
 
    # All checks passed — update last-run date
    _save_last_run_date(today)
    return {'tampered': False, 'reason': '', 'source': 'ok'}
 
 
# ─── System Code (machine fingerprint) ────────────────────────────────────────
def get_system_code() -> str:
    """Build a stable machine fingerprint from hardware identifiers."""
    parts = []
 
    try:
        parts.append(str(uuid.getnode()))
    except Exception:
        pass
    try:
        parts.append(platform.node())
    except Exception:
        pass
    try:
        parts.append(platform.machine())
    except Exception:
        pass
    try:
        parts.append(platform.processor())
    except Exception:
        pass
 
    for path in ['/etc/machine-id', '/var/lib/dbus/machine-id']:
        try:
            with open(path) as f:
                parts.append(f.read().strip())
            break
        except Exception:
            pass
    if platform.system() == 'Windows':
        try:
            import winreg
            key = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                 r'SOFTWARE\Microsoft\Cryptography')
            val, _ = winreg.QueryValueEx(key, 'MachineGuid')
            parts.append(val)
        except Exception:
            pass
 
    raw    = '|'.join(parts)
    digest = hashlib.sha256(raw.encode()).hexdigest()
    d      = digest[:16].upper()
    return '-'.join([d[i:i+4] for i in range(0, 16, 4)])
 
 
# ─── License validation ───────────────────────────────────────────────────────
def _load_public_key():
    return serialization.load_pem_public_key(PUBLIC_KEY_PEM, backend=default_backend())
 
def verify_license_key(license_key: str, system_code: str) -> dict:
    """
    Returns dict with keys:
      valid (bool), expired (bool), expiry (str), customer (str),
      issued (str), days_left (int), error (str), clock_tampered (bool)
    """
    try:
        # Sanitize
        clean_key = license_key.strip()
        clean_key = clean_key.replace('\r', '').replace('\n', '').replace(' ', '')
        clean_key = clean_key.lstrip('\ufeff')
 
        raw         = base64.b64decode(clean_key.encode())
        bundle      = json.loads(raw)
        payload_b64 = bundle['payload']
        sig_b64     = bundle['signature']
 
        payload_bytes = base64.b64decode(payload_b64)
        sig_bytes     = base64.b64decode(sig_b64)
 
        # 1. Verify RSA signature
        pub_key = _load_public_key()
        pub_key.verify(sig_bytes, payload_bytes, padding.PKCS1v15(), hashes.SHA256())
 
        # 2. Decode payload
        data = json.loads(payload_bytes)
 
        # 3. Check system code binding
        if data.get('system_code', '').upper() != system_code.upper():
            return {'valid': False, 'error': 'License is not valid for this machine.', 'clock_tampered': False}
 
        # 4. Check product
        if data.get('product', '') != 'HBILLSOFT':
            return {'valid': False, 'error': 'Invalid product in license.', 'clock_tampered': False}
 
        # 5. Clock tamper detection (all three layers)
        issued_str  = data.get('issued', datetime.date.today().isoformat())
        tamper_info = detect_clock_tampering(issued_str)
        if tamper_info['tampered']:
            return {
                'valid':          False,
                'expired':        False,
                'clock_tampered': True,
                'error':          tamper_info['reason'],
                'expiry':         data.get('expiry', ''),
                'customer':       data.get('customer', ''),
                'issued':         issued_str,
                'days_left':      0,
            }
 
        # 6. Check expiry
        expiry_date = datetime.date.fromisoformat(data['expiry'])
        today       = datetime.date.today()
        expired     = today > expiry_date
        days_left   = (expiry_date - today).days if not expired else 0
 
        return {
            'valid':          True,
            'expired':        expired,
            'expiry':         data['expiry'],
            'customer':       data.get('customer', ''),
            'issued':         issued_str,
            'days_left':      days_left,
            'clock_tampered': False,
            'error':          ''
        }
    except Exception as e:
        return {'valid': False, 'error': f'License verification failed: {e}', 'clock_tampered': False}
 
 
def load_saved_license() -> str:
    try:
        with open(LICENSE_FILE, 'r') as f:
            return f.read().strip()
    except Exception:
        return ''
 
def save_license(key: str):
    with open(LICENSE_FILE, 'w') as f:
        f.write(key.strip())
 
# ─── Config helpers ───────────────────────────────────────────────────────────
def load_config_from_file():
    try:
        with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}
 
def save_config_to_file(data):
    try:
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return True
    except Exception:
        return False
 
# ─── SQLite Database Functions ────────────────────────────────────────────────
def init_database():
    """Initialize SQLite database with orders, items, menu and categories tables."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Create orders table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS orders (
                id TEXT PRIMARY KEY,
                date TEXT NOT NULL,
                subtotal REAL NOT NULL,
                discount REAL NOT NULL,
                sgst REAL NOT NULL,
                cgst REAL NOT NULL,
                total REAL NOT NULL,
                payment_method TEXT NOT NULL DEFAULT 'Cash',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        # Migrate: add payment_method column if upgrading from older DB
        try:
            cursor.execute("ALTER TABLE orders ADD COLUMN payment_method TEXT NOT NULL DEFAULT 'Cash'")
        except Exception:
            pass  # Column already exists
        
        # Create order_items table (links items to orders)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS order_items (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_id TEXT NOT NULL,
                item_name TEXT NOT NULL,
                item_id TEXT,
                quantity INTEGER NOT NULL,
                price REAL NOT NULL,
                item_total REAL NOT NULL,
                FOREIGN KEY(order_id) REFERENCES orders(id)
            )
        ''')
 
        # Create categories table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS categories (
                id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                icon TEXT NOT NULL DEFAULT '🍽️',
                sort_order INTEGER DEFAULT 0
            )
        ''')
 
        # Create menu_items table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS menu_items (
                id INTEGER PRIMARY KEY,
                name TEXT NOT NULL,
                price REAL NOT NULL,
                category TEXT NOT NULL,
                image TEXT DEFAULT '🍽️',
                image_data TEXT,
                sort_order INTEGER DEFAULT 0
            )
        ''')
 
        # Create settings table (single JSON row)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS app_settings (
                id   INTEGER PRIMARY KEY CHECK (id = 1),
                data TEXT NOT NULL
            )
        ''')
 
        # Create cart table (persists current in-progress cart)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cart (
                id   INTEGER PRIMARY KEY CHECK (id = 1),
                data TEXT NOT NULL
            )
        ''')
 
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"Database init error: {e}")
        return False
 
# ─── Menu & Categories DB helpers ─────────────────────────────────────────────
 
def save_categories_to_db(categories: list) -> dict:
    """Replace all categories in the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM categories')
        for i, cat in enumerate(categories):
            cursor.execute('''
                INSERT OR REPLACE INTO categories (id, name, icon, sort_order)
                VALUES (?, ?, ?, ?)
            ''', (cat.get('id', ''), cat.get('name', ''), cat.get('icon', '🍽️'), i))
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}
 
def load_categories_from_db() -> dict:
    """Load all categories from the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT id, name, icon FROM categories ORDER BY sort_order, rowid')
        rows = cursor.fetchall()
        conn.close()
        categories = [{'id': r['id'], 'name': r['name'], 'icon': r['icon']} for r in rows]
        return {'ok': True, 'categories': categories}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'categories': []}
 
def save_menu_to_db(menu: list) -> dict:
    """Replace all menu items in the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM menu_items')
        for i, item in enumerate(menu):
            cursor.execute('''
                INSERT OR REPLACE INTO menu_items (id, name, price, category, image, image_data, sort_order)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                item.get('id'),
                item.get('name', ''),
                float(item.get('price', 0)),
                item.get('category', ''),
                item.get('image', '🍽️'),
                item.get('imageData', None),
                i
            ))
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}
 
def load_menu_from_db() -> dict:
    """Load all menu items from the DB."""
    try:
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT id, name, price, category, image, image_data FROM menu_items ORDER BY sort_order, id')
        rows = cursor.fetchall()
        conn.close()
        menu = []
        for r in rows:
            item = {
                'id': r['id'],
                'name': r['name'],
                'price': r['price'],
                'category': r['category'],
                'image': r['image']
            }
            if r['image_data']:
                item['imageData'] = r['image_data']
            menu.append(item)
        return {'ok': True, 'menu': menu}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'menu': []}
 
def save_settings_to_db(settings: dict) -> dict:
    """Persist app settings as a single JSON blob."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        # Strip internal-only keys before saving
        clean = {k: v for k, v in settings.items() if k != 'sessionCount'}
        cursor.execute(
            'INSERT OR REPLACE INTO app_settings (id, data) VALUES (1, ?)',
            (json.dumps(clean, ensure_ascii=False),)
        )
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}
 
def load_settings_from_db() -> dict:
    """Load app settings from SQLite. Returns {} if not yet saved."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT data FROM app_settings WHERE id = 1')
        row = cursor.fetchone()
        conn.close()
        if row:
            return {'ok': True, 'settings': json.loads(row[0])}
        return {'ok': True, 'settings': {}}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'settings': {}}
 
def save_cart_to_db(cart: list) -> dict:
    """Persist the current cart as a single JSON blob."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute(
            'INSERT OR REPLACE INTO cart (id, data) VALUES (1, ?)',
            (json.dumps(cart, ensure_ascii=False),)
        )
        conn.commit()
        conn.close()
        return {'ok': True}
    except Exception as e:
        return {'ok': False, 'error': str(e)}
 
def load_cart_from_db() -> dict:
    """Load the persisted cart from SQLite. Returns [] if empty."""
    try:
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        cursor.execute('SELECT data FROM cart WHERE id = 1')
        row = cursor.fetchone()
        conn.close()
        if row:
            return {'ok': True, 'cart': json.loads(row[0])}
        return {'ok': True, 'cart': []}
    except Exception as e:
        return {'ok': False, 'error': str(e), 'cart': []}
 
def save_order_to_db(order: dict) -> dict:
    """Save an order and its items to SQLite database."""
    try:
        init_database()
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        order_id       = order.get('id', str(uuid.uuid4()))
        date           = order.get('date', datetime.datetime.now().isoformat())
        subtotal       = float(order.get('subtotal', 0))
        discount       = float(order.get('discount', 0))
        sgst           = float(order.get('sgst', 0))
        cgst           = float(order.get('cgst', 0))
        total          = float(order.get('total', 0))
        payment_method = order.get('paymentMethod', 'Cash')

        # Save order
        cursor.execute('''
            INSERT OR REPLACE INTO orders
            (id, date, subtotal, discount, sgst, cgst, total, payment_method)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (order_id, date, subtotal, discount, sgst, cgst, total, payment_method))
        
        # Save order items
        for item in order.get('items', []):
            cursor.execute('''
                INSERT INTO order_items 
                (order_id, item_name, item_id, quantity, price, item_total)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (
                order_id,
                item.get('name', ''),
                item.get('id', ''),
                item.get('qty', 1),
                item.get('price', 0),
                float(item.get('qty', 1)) * float(item.get('price', 0))
            ))
        
        conn.commit()
        conn.close()
        return {'ok': True, 'order_id': order_id}
    except Exception as e:
        print(f"Error saving order to DB: {e}")
        return {'ok': False, 'error': str(e)}
 
def load_orders_from_db() -> dict:
    """Load all orders from SQLite database."""
    try:
        init_database()
        if not os.path.exists(DB_FILE):
            return {'ok': True, 'orders': []}
        
        conn = sqlite3.connect(DB_FILE)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Get all orders
        cursor.execute('SELECT * FROM orders ORDER BY date DESC')
        orders_rows = cursor.fetchall()
        
        orders = []
        for order_row in orders_rows:
            order_id = order_row['id']
            
            # Get items for this order
            cursor.execute('''
                SELECT item_name, item_id, quantity, price, item_total 
                FROM order_items 
                WHERE order_id = ?
            ''', (order_id,))
            items_rows = cursor.fetchall()
            
            items = [
                {
                    'name': item['item_name'],
                    'id': item['item_id'],
                    'qty': item['quantity'],
                    'price': item['price'],
                    'total': item['item_total']
                }
                for item in items_rows
            ]
            
            order = {
                'id': order_row['id'],
                'date': order_row['date'],
                'items': items,
                'subtotal': order_row['subtotal'],
                'discount': order_row['discount'],
                'sgst': order_row['sgst'],
                'cgst': order_row['cgst'],
                'gst': order_row['sgst'],  # For compatibility
                'total': order_row['total']
            }
            orders.append(order)
        
        conn.close()
        return {'ok': True, 'orders': orders}
    except Exception as e:
        print(f"Error loading orders from DB: {e}")
        return {'ok': False, 'error': str(e), 'orders': []}
 
def get_sales_summary(from_date=None, to_date=None) -> dict:
    """
    Get sales summary with items grouped by name, quantities, and date-based filtering.
    Returns items with total quantity sold and total revenue.
    """
    try:
        init_database()
        if not os.path.exists(DB_FILE):
            return {'ok': True, 'summary': [], 'total_revenue': 0}
        
        conn = sqlite3.connect(DB_FILE)
        cursor = conn.cursor()
        
        # Build query with optional date filtering
        query = '''
            SELECT 
                oi.item_name,
                SUM(oi.quantity) as total_qty,
                AVG(oi.price) as avg_price,
                SUM(oi.item_total) as total_revenue,
                COUNT(DISTINCT oi.order_id) as order_count
            FROM order_items oi
            JOIN orders o ON oi.order_id = o.id
        '''
        
        params = []
        if from_date or to_date:
            query += ' WHERE'
            if from_date:
                query += ' o.date >= ?'
                params.append(from_date)
            if from_date and to_date:
                query += ' AND'
            if to_date:
                query += ' o.date <= ?'
                params.append(to_date)
        
        query += ' GROUP BY oi.item_name ORDER BY total_revenue DESC'
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        summary = []
        total_revenue = 0
        for row in rows:
            item_summary = {
                'name': row[0],
                'quantity': int(row[1]),
                'avg_price': round(row[2], 2),
                'revenue': round(row[3], 2),
                'orders': int(row[4])
            }
            summary.append(item_summary)
            total_revenue += row[3]
        # Also compute total distinct orders, tax totals, and payment breakdown
        orders_count = None
        total_sgst = 0.0
        total_cgst = 0.0
        grand_total_including_tax = 0.0
        payment_breakdown = {'Cash': 0.0, 'Card': 0.0, 'UPI': 0.0, 'Other': 0.0}
        try:
            count_query = 'SELECT COUNT(DISTINCT o.id) FROM orders o'
            sum_query = (
                "SELECT SUM(o.sgst), SUM(o.cgst), SUM(o.total),"
                " SUM(CASE WHEN o.payment_method='Cash'  THEN o.total ELSE 0 END),"
                " SUM(CASE WHEN o.payment_method='Card'  THEN o.total ELSE 0 END),"
                " SUM(CASE WHEN o.payment_method='UPI'   THEN o.total ELSE 0 END),"
                " SUM(CASE WHEN o.payment_method NOT IN ('Cash','Card','UPI') THEN o.total ELSE 0 END)"
                " FROM orders o"
            )
            if from_date or to_date:
                where_parts = []
                params_count = []
                if from_date:
                    where_parts.append('o.date >= ?')
                    params_count.append(from_date)
                if to_date:
                    where_parts.append('o.date <= ?')
                    params_count.append(to_date)
                where_clause = ' WHERE ' + ' AND '.join(where_parts)
                cursor.execute(count_query + where_clause, params_count)
                orders_count = cursor.fetchone()[0]
                cursor.execute(sum_query + where_clause, params_count)
            else:
                cursor.execute(count_query)
                orders_count = cursor.fetchone()[0]
                cursor.execute(sum_query)

            sums = cursor.fetchone()
            if sums:
                total_sgst = float(sums[0] or 0)
                total_cgst = float(sums[1] or 0)
                grand_total_including_tax = float(sums[2] or 0)
                payment_breakdown['Cash']  = round(float(sums[3] or 0), 2)
                payment_breakdown['Card']  = round(float(sums[4] or 0), 2)
                payment_breakdown['UPI']   = round(float(sums[5] or 0), 2)
                payment_breakdown['Other'] = round(float(sums[6] or 0), 2)
        except Exception:
            orders_count = None

        conn.close()
        return {
            'ok': True,
            'summary': summary,
            'total_revenue': round(total_revenue, 2),
            'orders_count': orders_count,
            'total_sgst': round(total_sgst, 2),
            'total_cgst': round(total_cgst, 2),
            'total_gst': round(total_sgst + total_cgst, 2),
            'grand_total': round(grand_total_including_tax, 2),
            'payment_breakdown': payment_breakdown,
        }
    except Exception as e:
        print(f"Error getting sales summary: {e}")
        return {'ok': False, 'error': str(e), 'summary': [], 'total_revenue': 0}
 
# ─── pywebview API ────────────────────────────────────────────────────────────
class Api:
    def close_window(self):
        window.destroy()
 
    def load_config(self):
        return load_config_from_file()
 
    def save_config(self, data):
        return save_config_to_file(data)
 
    def get_session_count(self):
        return load_config_from_file().get('sessionCount', 0)
 
    # --- License API ---
    def get_system_code(self):
        return get_system_code()
 
    def activate_license(self, license_key: str):
        sc     = get_system_code()
        result = verify_license_key(license_key, sc)
        if result['valid'] and not result['expired'] and not result.get('clock_tampered'):
            save_license(license_key)
        return result
 
    def check_license(self):
        sc  = get_system_code()
        key = load_saved_license()
        if not key:
            return {'valid': False, 'expired': False, 'error': 'No license found.', 'system_code': sc, 'clock_tampered': False}
        result = verify_license_key(key, sc)
        result['system_code'] = sc
        return result
 
    def renew_license(self, new_key: str):
        """Renew/replace an expired license with a new key."""
        return self.activate_license(new_key)
 
    # --- Mobile Server API ---
    def get_mobile_server_info(self):
        try:
            import mobile_server
            return {
                'ok': True,
                'url': mobile_server.MOBILE_URL,
                'ip': mobile_server.LOCAL_IP,
                'port': mobile_server.MOBILE_PORT
            }
        except Exception as e:
            return {'ok': False, 'error': str(e)}
 
    def get_pending_mobile_orders(self):
        try:
            import mobile_server
            return {'ok': True, 'orders': mobile_server.get_all_pending()}
        except Exception as e:
            return {'ok': False, 'error': str(e)}
 
    def dismiss_mobile_order(self, order_id):
        try:
            import mobile_server
            ok = mobile_server.delete_pending(order_id)
            return {'ok': ok}
        except Exception as e:
            return {'ok': False, 'error': str(e)}
 
    # --- Excel Sales DB ---
    def get_excel_path(self) -> str:
        """Return the path of the auto-saved sales Excel file."""
        return EXCEL_DB_FILE
 
    def save_order_to_excel(self, order: dict) -> dict:
        """
        After each completed order, rebuild the monthly sheet in sales_data.xlsx.
        Format: one sheet per month (e.g. "May 2026"), day-by-day sections.
        Each day: date header, column headers (#/Item Name/Category/Qty/Rate/Amount),
        item rows, Subtotal, SGST, CGST, Day Total.
        """
        if not _OPENPYXL_OK:
            return {'ok': False, 'error': 'openpyxl not installed. Run: pip install openpyxl'}

        try:
            r = load_settings_from_db()
            cfg        = r['settings'] if r['ok'] and r['settings'] else load_config_from_file()
            restaurant = cfg.get('restaurantName', 'MY RESTAURANT')
            sgst_rate  = float(cfg.get('sgst', 2.5))
            cgst_rate  = float(cfg.get('cgst', 2.5))

            order_dt    = datetime.datetime.fromisoformat(order.get('date', datetime.datetime.now().isoformat()))
            sheet_name  = order_dt.strftime('%B %Y')
            import calendar
            month_start  = order_dt.strftime('%Y-%m-01') + 'T00:00:00'
            last_day     = calendar.monthrange(order_dt.year, order_dt.month)[1]
            month_end    = order_dt.strftime(f'%Y-%m-{last_day:02d}') + 'T23:59:59'

            conn = sqlite3.connect(DB_FILE)
            conn.row_factory = sqlite3.Row
            cur  = conn.cursor()

            cur.execute('''
                SELECT
                    date(o.date)               AS day,
                    oi.item_name,
                    COALESCE(mi.category, '')  AS category,
                    SUM(oi.quantity)            AS qty,
                    oi.price                    AS rate,
                    SUM(oi.item_total)           AS amount
                FROM order_items oi
                JOIN orders o        ON oi.order_id = o.id
                LEFT JOIN menu_items mi ON oi.item_id = mi.id
                WHERE o.date >= ? AND o.date <= ?
                GROUP BY date(o.date), oi.item_name, oi.price
                ORDER BY date(o.date), oi.item_name
            ''', [month_start, month_end])
            rows = cur.fetchall()

            cur.execute('''
                SELECT date(date) AS day,
                       SUM(subtotal)  AS subtotal,
                       SUM(sgst)      AS sgst,
                       SUM(cgst)      AS cgst,
                       SUM(total)     AS total,
                       SUM(CASE WHEN payment_method = 'Cash'   THEN total ELSE 0 END) AS cash_total,
                       SUM(CASE WHEN payment_method = 'Card'   THEN total ELSE 0 END) AS card_total,
                       SUM(CASE WHEN payment_method = 'UPI'    THEN total ELSE 0 END) AS upi_total,
                       SUM(CASE WHEN payment_method NOT IN ('Cash','Card','UPI') THEN total ELSE 0 END) AS other_total
                FROM orders
                WHERE date >= ? AND date <= ?
                GROUP BY date(date)
                ORDER BY date(date)
            ''', [month_start, month_end])
            day_tax = {r['day']: dict(r) for r in cur.fetchall()}
            conn.close()

            from collections import OrderedDict
            days = OrderedDict()
            for row in rows:
                d = row['day']
                if d not in days:
                    days[d] = []
                days[d].append(dict(row))

            # ── Style helpers ────────────────────────────────────────────
            HDR_FILL  = PatternFill('solid', fgColor='1a3c5e')
            SUB_FILL  = PatternFill('solid', fgColor='EAF2FB')
            DAY_FILL  = PatternFill('solid', fgColor='1d5c2e')
            DATE_FILL = PatternFill('solid', fgColor='DFF0D8')
            ODD_FILL  = PatternFill('solid', fgColor='F7F9FC')
            EVEN_FILL = PatternFill('solid', fgColor='FFFFFF')
            TAX_FILL  = PatternFill('solid', fgColor='FFF9E6')

            def mk_border(style='thin', color='BBBBBB'):
                s = Side(style=style, color=color)
                return Border(left=s, right=s, top=s, bottom=s)

            def thin():   return mk_border()
            def thick():  return mk_border('medium', '1a3c5e')

            c_center = Alignment(horizontal='center', vertical='center')
            c_left   = Alignment(horizontal='left',   vertical='center', indent=1)
            c_right  = Alignment(horizontal='right',  vertical='center')

            def wc(ws, row, col, value, font=None, fill=None, align=None, border=None, nfmt=None):
                cell = ws.cell(row=row, column=col, value=value)
                if font:  cell.font            = font
                if fill:  cell.fill            = fill
                if align: cell.alignment       = align
                if border: cell.border         = border
                if nfmt:  cell.number_format   = nfmt
                return cell

            # ── Load or create workbook ──────────────────────────────────
            if os.path.exists(EXCEL_DB_FILE):
                wb = load_workbook(EXCEL_DB_FILE)
            else:
                wb = Workbook()
                if 'Sheet' in wb.sheetnames:
                    del wb['Sheet']

            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(title=sheet_name)

            ws.column_dimensions['A'].width = 5
            ws.column_dimensions['B'].width = 28
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 7
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 16

            # Row 1: Restaurant name
            ws.merge_cells('A1:F1')
            wc(ws, 1, 1, restaurant.upper(),
               font=Font(bold=True, size=15, color='FFFFFF'),
               fill=HDR_FILL, align=c_center)
            ws.row_dimensions[1].height = 26

            # Row 2: Report title
            ws.merge_cells('A2:F2')
            wc(ws, 2, 1, f'Monthly Sales Report  —  {sheet_name}',
               font=Font(bold=True, size=11, color='FFFFFF'),
               fill=HDR_FILL, align=c_center)
            ws.row_dimensions[2].height = 18

            ws.row_dimensions[3].height = 8
            cur_row = 4

            for day_str, items in days.items():
                dt      = datetime.datetime.strptime(day_str, '%Y-%m-%d')
                day_lbl = dt.strftime('%A, %d %B %Y')
                taxes   = day_tax.get(day_str, {})
                day_sub   = sum(float(it['amount']) for it in items)
                day_sgst  = float(taxes.get('sgst') or round(day_sub * sgst_rate / (100 + sgst_rate + cgst_rate), 2))
                day_cgst  = float(taxes.get('cgst') or round(day_sub * cgst_rate / (100 + sgst_rate + cgst_rate), 2))
                day_total = float(taxes.get('total') or day_sub)
                pay_cash  = float(taxes.get('cash_total') or 0)
                pay_card  = float(taxes.get('card_total') or 0)
                pay_upi   = float(taxes.get('upi_total')  or 0)
                pay_other = float(taxes.get('other_total') or 0)

                # Date label
                ws.merge_cells(f'A{cur_row}:F{cur_row}')
                wc(ws, cur_row, 1, f'  {day_lbl}',
                   font=Font(bold=True, size=10, color='1d5c2e'),
                   fill=DATE_FILL,
                   align=Alignment(horizontal='left', vertical='center', indent=1))
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1

                # Column headers
                headers = ['#', 'Item Name', 'Category', 'Qty', 'Rate (Rs.)', 'Amount (Rs.)']
                for ci, h in enumerate(headers, 1):
                    wc(ws, cur_row, ci, h,
                       font=Font(bold=True, size=9, color='FFFFFF'),
                       fill=HDR_FILL, align=c_center, border=thin())
                ws.row_dimensions[cur_row].height = 16
                cur_row += 1

                # Item rows
                for idx, it in enumerate(items):
                    fill = ODD_FILL if idx % 2 == 0 else EVEN_FILL
                    cat = str(it['category']).replace('-', ' ').title() if it['category'] else ''
                    vals   = [idx+1, it['item_name'], cat, int(it['qty']), round(float(it['rate']),2), round(float(it['amount']),2)]
                    aligns = [c_center, c_left, c_left, c_center, c_right, c_right]
                    for ci, (v, al) in enumerate(zip(vals, aligns), 1):
                        cell = wc(ws, cur_row, ci, v, font=Font(size=9), fill=fill, align=al, border=thin())
                        if ci in (5, 6):
                            cell.number_format = '#,##0.00'
                    ws.row_dimensions[cur_row].height = 15
                    cur_row += 1

                # Subtotal
                ws.merge_cells(f'A{cur_row}:E{cur_row}')
                wc(ws, cur_row, 1, 'Subtotal', font=Font(bold=True, size=9),
                   fill=SUB_FILL, align=c_left, border=thin())
                wc(ws, cur_row, 6, round(day_sub, 2), font=Font(bold=True, size=9),
                   fill=SUB_FILL, align=c_right, border=thin(), nfmt='#,##0.00')
                ws.row_dimensions[cur_row].height = 15
                cur_row += 1

                # SGST
                ws.merge_cells(f'A{cur_row}:E{cur_row}')
                wc(ws, cur_row, 1, f'SGST  ({sgst_rate}%)', font=Font(size=9),
                   fill=TAX_FILL, align=c_left, border=thin())
                wc(ws, cur_row, 6, round(day_sgst, 2), font=Font(size=9),
                   fill=TAX_FILL, align=c_right, border=thin(), nfmt='#,##0.00')
                ws.row_dimensions[cur_row].height = 14
                cur_row += 1

                # CGST
                ws.merge_cells(f'A{cur_row}:E{cur_row}')
                wc(ws, cur_row, 1, f'CGST  ({cgst_rate}%)', font=Font(size=9),
                   fill=TAX_FILL, align=c_left, border=thin())
                wc(ws, cur_row, 6, round(day_cgst, 2), font=Font(size=9),
                   fill=TAX_FILL, align=c_right, border=thin(), nfmt='#,##0.00')
                ws.row_dimensions[cur_row].height = 14
                cur_row += 1

                # Day Total
                ws.merge_cells(f'A{cur_row}:E{cur_row}')
                wc(ws, cur_row, 1, f'Day Total  —  {day_lbl}',
                   font=Font(bold=True, size=10, color='FFFFFF'),
                   fill=DAY_FILL, align=c_left, border=thick())
                wc(ws, cur_row, 6, round(day_total, 2),
                   font=Font(bold=True, size=10, color='FFFFFF'),
                   fill=DAY_FILL, align=c_right, border=thick(), nfmt='#,##0.00')
                ws.row_dimensions[cur_row].height = 18
                cur_row += 1

                # Payment method breakdown
                PAY_FILL = PatternFill('solid', fgColor='EAF2FB')
                for label, amount in [('Cash', pay_cash), ('Card', pay_card),
                                       ('UPI', pay_upi), ('Other', pay_other)]:
                    if amount > 0:
                        ws.merge_cells(f'A{cur_row}:E{cur_row}')
                        wc(ws, cur_row, 1, f'  └ {label}',
                           font=Font(size=9, italic=True), fill=PAY_FILL,
                           align=c_left, border=thin())
                        wc(ws, cur_row, 6, round(amount, 2),
                           font=Font(size=9, italic=True), fill=PAY_FILL,
                           align=c_right, border=thin(), nfmt='#,##0.00')
                        ws.row_dimensions[cur_row].height = 14
                        cur_row += 1

                # Gap between days
                ws.row_dimensions[cur_row].height = 8
                cur_row += 1

            wb.save(EXCEL_DB_FILE)
            return {'ok': True, 'file': EXCEL_DB_FILE, 'sheet': sheet_name}

        except Exception as e:
            import traceback
            return {'ok': False, 'error': str(e) + '\n' + traceback.format_exc()}


    def load_orders_from_excel(self) -> dict:
        """
        Load all orders from the persistent sales_data.xlsx file.
        Returns {'ok': True, 'orders': [...]} or {'ok': False, 'error': '...', 'orders': []}.
        """
        if not _OPENPYXL_OK:
            return {'ok': False, 'error': 'openpyxl not installed', 'orders': []}
 
        orders = []
        try:
            if not os.path.exists(EXCEL_DB_FILE):
                return {'ok': True, 'orders': []}
 
            wb = load_workbook(EXCEL_DB_FILE)
            
            # Load orders from all sheets (one sheet per day)
            for sheet_name in wb.sheetnames:
                try:
                    ws = wb[sheet_name]
                    # Skip non-data rows (banner, header, etc.)
                    # Data starts at row 5
                    for row_idx, row in enumerate(ws.iter_rows(min_row=5, values_only=True), start=5):
                        if not row[0] or row[0] == 'Date':  # Skip empty/header rows
                            continue
                        
                        try:
                            # Parse date and time
                            date_str = str(row[0]) if row[0] else ''
                            time_str = str(row[1]) if row[1] else '00:00 AM'
                            datetime_str = f"{date_str} {time_str}"
                            
                            # Parse datetime - try multiple formats
                            order_dt = None
                            for fmt in ['%d-%b-%Y %I:%M %p', '%Y-%m-%d %H:%M:%S', '%d-%b-%Y %H:%M:%S']:
                                try:
                                    order_dt = datetime.datetime.strptime(datetime_str, fmt)
                                    break
                                except:
                                    continue
                            
                            if not order_dt:
                                order_dt = datetime.datetime.now()
                            
                            # Parse amounts
                            subtotal = float(row[4] or 0)
                            discount = float(row[5] or 0)
                            sgst = float(row[6] or 0)
                            cgst = float(row[7] or 0)
                            total = float(row[8] or 0)
                            
                            order = {
                                'id': str(row[2]) if row[2] else '',
                                'date': order_dt.isoformat(),
                                'items': [],  # Excel doesn't store full item details, only summary
                                'subtotal': subtotal,
                                'discount': discount,
                                'sgst': sgst,
                                'cgst': cgst,
                                'gst': sgst,  # For compatibility
                                'total': total
                            }
                            orders.append(order)
                        except Exception as row_err:
                            print(f"Error parsing row {row_idx} in sheet '{sheet_name}': {row_err}")
                            continue
                except Exception as sheet_err:
                    print(f"Error reading sheet '{sheet_name}': {sheet_err}")
                    continue
            
            return {'ok': True, 'orders': orders}
 
        except Exception as e:
            return {'ok': False, 'error': str(e), 'orders': []}
 
    def save_order(self, order: dict) -> dict:
        """Save order to SQLite database."""
        return save_order_to_db(order)
 
    def load_orders(self) -> dict:
        """Load all orders from SQLite database."""
        return load_orders_from_db()
 
    # --- Settings SQLite API ---
    def save_settings(self, settings: dict) -> dict:
        """Persist settings to SQLite (and config.json for legacy compat)."""
        save_config_to_file(settings)          # keep config.json in sync
        return save_settings_to_db(settings)
 
    def load_settings(self) -> dict:
        """Load settings from SQLite; fall back to config.json."""
        result = load_settings_from_db()
        if result['ok'] and result['settings']:
            return result
        # Fall back to config.json (first-run or migration)
        r = load_settings_from_db(); cfg = r['settings'] if r['ok'] and r['settings'] else load_config_from_file()
        cfg.pop('sessionCount', None)
        return {'ok': True, 'settings': cfg}
 
    # --- Cart SQLite API ---
    def save_cart(self, cart: list) -> dict:
        """Persist cart to SQLite."""
        return save_cart_to_db(cart)
 
    def load_cart(self) -> dict:
        """Load cart from SQLite."""
        return load_cart_from_db()
 
    # --- Menu & Categories SQLite API ---
    def save_menu(self, menu: list) -> dict:
        """Persist full menu to SQLite."""
        return save_menu_to_db(menu)
 
    def load_menu(self) -> dict:
        """Load full menu from SQLite."""
        return load_menu_from_db()
 
    def save_categories(self, categories: list) -> dict:
        """Persist full categories list to SQLite."""
        return save_categories_to_db(categories)
 
    def load_categories(self) -> dict:
        """Load full categories list from SQLite."""
        return load_categories_from_db()
 
    def get_sales_summary_api(self, from_date=None, to_date=None) -> dict:
        """Get sales summary with items and revenue."""
        return get_sales_summary(from_date, to_date)
 
    def print_kot_win32(self, kot_text: str) -> dict:
        """
        Send plain-text KOT to the default Windows printer using win32print.
        Falls back gracefully if win32print is not available (Linux/macOS).
        """
        try:
            import win32print
            printer_name = win32print.GetDefaultPrinter()
            hprinter = win32print.OpenPrinter(printer_name)
            try:
                hjob = win32print.StartDocPrinter(hprinter, 1, ("HBILLSOFT KOT", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hprinter)
                    data = (kot_text + '\n\n\n\n').encode('cp437', errors='replace')
                    win32print.WritePrinter(hprinter, data)
                    win32print.EndPagePrinter(hprinter)
                finally:
                    win32print.EndDocPrinter(hprinter)
            finally:
                win32print.ClosePrinter(hprinter)
            return {'ok': True, 'printer': printer_name}
        except ImportError:
            return {'ok': False, 'error': 'win32print not available'}
        except Exception as e:
            return {'ok': False, 'error': str(e)}

    def print_receipt_win32(self, receipt_text: str) -> dict:
        """
        Send plain-text receipt to the default Windows printer using win32print.
        Falls back gracefully if win32print is not available (Linux/macOS).
        """
        try:
            import win32print
            import win32ui
            from PIL import ImageWin   # not needed for raw text printing
 
            printer_name = win32print.GetDefaultPrinter()
 
            # Open printer and send a raw ESC/POS-compatible text job
            hprinter = win32print.OpenPrinter(printer_name)
            try:
                hjob = win32print.StartDocPrinter(hprinter, 1, ("HBILLSOFT Receipt", None, "RAW"))
                try:
                    win32print.StartPagePrinter(hprinter)
                    # Encode as CP437 (standard for most thermal/receipt printers)
                    data = (receipt_text + '\n\n\n\n').encode('cp437', errors='replace')
                    win32print.WritePrinter(hprinter, data)
                    win32print.EndPagePrinter(hprinter)
                finally:
                    win32print.EndDocPrinter(hprinter)
            finally:
                win32print.ClosePrinter(hprinter)
 
            return {'ok': True, 'printer': printer_name}
 
        except ImportError:
            # win32print not available (non-Windows) — tell JS to use window.print()
            return {'ok': False, 'error': 'win32print not available — use window.print() fallback'}
        except Exception as e:
            return {'ok': False, 'error': str(e)}

    def export_sales_excel(self, from_date=None, to_date=None) -> dict:
        """
        Build a clean sales report Excel file and save it to the desktop.
        Returns {'ok': True, 'file': path} or {'ok': False, 'error': '...'}.
        """
        if not _OPENPYXL_OK:
            return {'ok': False, 'error': 'openpyxl not installed. Run: pip install openpyxl'}
 
        try:
            r = load_settings_from_db(); cfg = r['settings'] if r['ok'] and r['settings'] else load_config_from_file()
            curr = cfg.get('currency', '₹')
            restaurant = cfg.get('restaurantName', 'HBILLSOFT')
 
            # ── Fetch aggregated data ───────────────────────────────────────
            summary = get_sales_summary(from_date, to_date)
            if not summary['ok']:
                return {'ok': False, 'error': summary.get('error', 'Failed to load data')}
 
            items       = summary.get('summary') or []      # [{name, quantity, avg_price, revenue}]
            sgst        = float(summary.get('total_sgst') or 0)
            cgst        = float(summary.get('total_cgst') or 0)
            grand_total = float(summary.get('grand_total') or 0)
            items_total = float(summary.get('total_revenue') or 0)  # pre-tax subtotal
            discount    = round(items_total + sgst + cgst - grand_total, 2)
            if discount < 0:
                discount = 0
 
            # ── Style helpers ───────────────────────────────────────────────
            BLUE       = '1a3c5e'
            WHITE      = 'FFFFFF'
            LIGHT_BLUE = 'EAF2FB'
            YELLOW     = 'FFF9C4'
            GREEN_DARK = '1d5c2e'
 
            def thin_border():
                s = Side(style='thin', color='BBBBBB')
                return Border(left=s, right=s, top=s, bottom=s)
 
            def thick_border():
                s = Side(style='medium', color=BLUE)
                return Border(left=s, right=s, top=s, bottom=s)
 
            center  = Alignment(horizontal='center', vertical='center')
            left    = Alignment(horizontal='left',   vertical='center', indent=1)
            right_a = Alignment(horizontal='right',  vertical='center')
 
            # ── Build workbook ──────────────────────────────────────────────
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sales Report'
 
            # Column widths: A=Item, B=Qty, C=Unit Price, D=Amount
            ws.column_dimensions['A'].width = 36
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
            ws.column_dimensions['D'].width = 18
 
            money_fmt = f'"{curr}"#,##0.00'
 
            # ── Header block (rows 1-3) ─────────────────────────────────────
            ws.merge_cells('A1:D1')
            c = ws['A1']
            c.value     = restaurant
            c.font      = Font(bold=True, size=15, color=WHITE)
            c.fill      = PatternFill('solid', fgColor=BLUE)
            c.alignment = center
            ws.row_dimensions[1].height = 24
 
            ws.merge_cells('A2:D2')
            c = ws['A2']
            c.value     = 'Sales Report'
            c.font      = Font(bold=True, size=11, color=WHITE)
            c.fill      = PatternFill('solid', fgColor=BLUE)
            c.alignment = center
            ws.row_dimensions[2].height = 18
 
            # Period label
            if from_date or to_date:
                fd = (from_date or '').split('T')[0]
                td = (to_date   or '').split('T')[0]
                period = f'Period: {fd or "Start"}  →  {td or "Today"}'
            else:
                period = f'Exported: {datetime.date.today().strftime("%d-%b-%Y")}'
 
            ws.merge_cells('A3:D3')
            c = ws['A3']
            c.value     = period
            c.font      = Font(italic=True, size=10, color='555555')
            c.fill      = PatternFill('solid', fgColor='F0F4F8')
            c.alignment = center
            ws.row_dimensions[3].height = 16
 
            ws.append([])   # row 4 blank
            ws.row_dimensions[4].height = 6
 
            # ── Column headers (row 5) ──────────────────────────────────────
            headers = ['Item Name', 'Qty Sold', f'Unit Price ({curr})', f'Amount ({curr})']
            ws.append(headers)
            for col_idx, _ in enumerate(headers, start=1):
                cell = ws.cell(row=5, column=col_idx)
                cell.font      = Font(bold=True, color=WHITE, size=10)
                cell.fill      = PatternFill('solid', fgColor=BLUE)
                cell.alignment = center
                cell.border    = thin_border()
            ws.row_dimensions[5].height = 18
 
            # ── Item rows ───────────────────────────────────────────────────
            for i, item in enumerate(items):
                row_num = 6 + i
                fill = PatternFill('solid', fgColor=LIGHT_BLUE) if i % 2 == 0 else PatternFill('solid', fgColor=WHITE)
                ws.append([
                    item['name'],
                    item['quantity'],
                    round(float(item['avg_price'] or 0), 2),
                    round(float(item['revenue'] or 0), 2),
                ])
                ws.cell(row=row_num, column=1).alignment = left
                ws.cell(row=row_num, column=2).alignment = center
                for col_idx in range(1, 5):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.fill   = fill
                    cell.border = thin_border()
                for col_idx in [3, 4]:
                    ws.cell(row=row_num, column=col_idx).number_format = money_fmt
 
            # ── Blank separator ─────────────────────────────────────────────
            sep_row = 6 + len(items)
            ws.append([])
            ws.row_dimensions[sep_row].height = 8
 
            # ── Totals block ────────────────────────────────────────────────
            total_qty = sum(it['quantity'] for it in items)
 
            totals = [
                ('Total Items Sold',          total_qty,                       '',        items_total),
                ('Discount',                  '',                              '',        discount),
                ('Subtotal (after discount)', '',                              '',        round(items_total - discount, 2)),
                ('SGST',                      '',                              '',        sgst),
                ('CGST',                      '',                              '',        cgst),
                ('Total GST',                 '',                              '',        round(sgst + cgst, 2)),
            ]
 
            for label, qty, _, amount in totals:
                r = ws.max_row + 1
                ws.append([label, qty, '', round(amount, 2)])
                ws.cell(row=r, column=1).font      = Font(size=10)
                ws.cell(row=r, column=1).alignment = left
                ws.cell(row=r, column=2).alignment = center
                ws.cell(row=r, column=4).number_format = money_fmt
                ws.cell(row=r, column=4).alignment = right_a
                for col_idx in range(1, 5):
                    ws.cell(row=r, column=col_idx).border = thin_border()
                    ws.cell(row=r, column=col_idx).fill   = PatternFill('solid', fgColor='F7F9FC')
 
            # ── Grand total row ─────────────────────────────────────────────
            ws.append([])   # blank row
            gt_row = ws.max_row + 1
            ws.append(['GRAND TOTAL', '', '', round(grand_total, 2)])
            ws.merge_cells(f'A{gt_row}:C{gt_row}')
            ws.cell(row=gt_row, column=1).value     = 'GRAND TOTAL'
            ws.cell(row=gt_row, column=1).font      = Font(bold=True, size=12, color=WHITE)
            ws.cell(row=gt_row, column=1).fill      = PatternFill('solid', fgColor=GREEN_DARK)
            ws.cell(row=gt_row, column=1).alignment = center
            ws.cell(row=gt_row, column=4).value          = round(grand_total, 2)
            ws.cell(row=gt_row, column=4).font           = Font(bold=True, size=12, color=WHITE)
            ws.cell(row=gt_row, column=4).fill           = PatternFill('solid', fgColor=GREEN_DARK)
            ws.cell(row=gt_row, column=4).number_format  = money_fmt
            ws.cell(row=gt_row, column=4).alignment      = right_a
            for col_idx in range(1, 5):
                ws.cell(row=gt_row, column=col_idx).border = thick_border()
            ws.row_dimensions[gt_row].height = 22
 
            # ── Save to Desktop ─────────────────────────────────────────────
            today_str = datetime.date.today().strftime('%Y-%m-%d')
            desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
            if not os.path.isdir(desktop):
                desktop = os.path.expanduser('~')
            out_path = os.path.join(desktop, f'HBILLSOFT_Sales_{today_str}.xlsx')
 
            wb.save(out_path)
            return {'ok': True, 'file': out_path}
 
        except Exception as e:
            return {'ok': False, 'error': str(e)}
 
# ─── Startup ──────────────────────────────────────────────────────────────────
api = Api()
 
# Initialize database
init_database()
 
config = load_config_from_file()
config['sessionCount'] = config.get('sessionCount', 0) + 1
save_config_to_file(config)
 
html_file = resource_path('RestoPOS.html')
 
# Start mobile server background thread
try:
    import mobile_server
    mobile_server.start_mobile_server()
except Exception as e:
    print(f"Failed to start mobile server: {e}")
 
window = webview.create_window(
    title='HBILLSOFT',
    url='file:///' + html_file.replace('\\', '/'),
    width=1280,
    height=800,
    min_size=(1024, 600),
    resizable=True,
    fullscreen=True,
    js_api=api
)
 
webview.start()